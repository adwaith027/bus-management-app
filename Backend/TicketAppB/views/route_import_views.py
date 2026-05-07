import io
import openpyxl
from openpyxl import Workbook
from django.db import transaction
from django.http import HttpResponse
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView

from ..models import BusType, Stage, Route, RouteStage, RouteBusType, Fare
from .auth_views import get_user_from_cookie


# ── Constants ─────────────────────────────────────────────────────────────────

FARE_TYPE_MAP = {'TABLE': 1, 'GRAPH': 2}

REQUIRED_COLS = {
    'routename', 'routecode', 'bustype', 'noofstages', 'minfare',
    'faretype', 'distance', 'half', 'student', 'adjust', 'luggage',
    'ph', 'concession', 'pass', 'stageorder', 'stagenames', 'fare',
}
FLAG_COLS = ['half', 'student', 'adjust', 'luggage', 'ph', 'concession', 'pass']
IGNORED_COLS = {'ladies', 'senior'}


# ── Auth helper ───────────────────────────────────────────────────────────────

def _auth(request):
    user = get_user_from_cookie(request)
    if not user:
        return None, None, Response({'message': 'Authentication required.'}, status=status.HTTP_401_UNAUTHORIZED)
    if user.role != 'company_admin':
        return None, None, Response({'message': 'Company admins only.'}, status=status.HTTP_403_FORBIDDEN)
    if not user.company:
        return None, None, Response({'message': 'No company mapped.'}, status=status.HTTP_400_BAD_REQUEST)
    return user, user.company, None


def _to_bool(val):
    try:
        return bool(int(float(str(val or 0))))
    except (ValueError, TypeError):
        return False


def _to_float(val):
    try:
        return float(str(val or '').strip())
    except (ValueError, TypeError):
        return None


# ── Core parser/validator ─────────────────────────────────────────────────────

def _load_worksheet(file_bytes):
    """Load openpyxl worksheet from bytes. Returns (ws, error_msg)."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return None, f'Cannot read file: {str(e)}'
    ws = wb['SRE_Import'] if 'SRE_Import' in wb.sheetnames else wb.active
    return ws, None


def _parse_and_validate(ws, company):
    """
    Parse worksheet, run all validation checks.
    Returns (route_groups, errors, warnings, duplicate_codes)
    route_groups: dict[route_code → {info dict, stages list}]
    errors: list[{route_code, row, message}]
    warnings: list[{message}]
    duplicate_codes: list[route_code strings]
    """
    errors = []
    warnings = []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return None, [{'route_code': None, 'row': None, 'message': 'File has no data rows.'}], [], []

    raw_headers = [str(h or '').strip() for h in rows[0]]
    headers_lower = [h.lower() for h in raw_headers]
    header_set = set(headers_lower)

    # A. Required columns
    missing = REQUIRED_COLS - header_set
    if missing:
        return None, [{'route_code': None, 'row': None,
                       'message': f'Missing required columns: {", ".join(sorted(missing))}'}], [], []

    # F. Incompatibility warnings
    ignored = IGNORED_COLS & header_set
    if ignored:
        cols_str = ', '.join(f'"{c.title()}"' for c in sorted(ignored))
        warnings.append({'message': f'Excel columns {cols_str} are not supported and will be ignored during import.'})

    # Locate column indices
    def col_idx(name):
        try:
            return headers_lower.index(name.lower())
        except ValueError:
            return -1

    fare_cidx = col_idx('fare')

    # Discover MatrixColK columns (for GRAPH fares)
    extra_cols = {}  # k → col_index (MatrixColK means fare from stage k to current stage)
    for i, h in enumerate(headers_lower):
        if h.startswith('matrixcol') and i > fare_cidx:
            try:
                k = int(h[len('matrixcol'):])
                extra_cols[k] = i
            except ValueError:
                pass

    def cell(row_values, name):
        idx = col_idx(name)
        val = row_values[idx] if 0 <= idx < len(row_values) else None
        return str(val or '').strip()

    def raw_cell(row_values, cidx):
        return row_values[cidx] if 0 <= cidx < len(row_values) else None

    # Pre-load bus types (keyed by bustype_code lowercase)
    bus_type_map = {bt.bustype_code.strip().lower(): bt
                    for bt in BusType.objects.filter(company=company, is_active=True)}

    route_groups = {}

    for row_num, row_values in enumerate(rows[1:], start=2):
        # Skip fully empty rows
        if all(v is None or str(v).strip() == '' for v in row_values):
            continue

        route_code = cell(row_values, 'routecode').upper()
        if not route_code:
            errors.append({'route_code': None, 'row': row_num, 'message': 'RouteCode is empty.'})
            continue

        route_name   = cell(row_values, 'routename')
        bus_type_raw = cell(row_values, 'bustype')
        fare_type_raw = cell(row_values, 'faretype').upper()
        stage_order_raw = cell(row_values, 'stageorder')
        stage_name   = cell(row_values, 'stagenames')

        # B. StageOrder
        try:
            stage_order = int(float(stage_order_raw)) if stage_order_raw else 0
            if stage_order <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            errors.append({'route_code': route_code, 'row': row_num,
                           'message': f'StageOrder "{stage_order_raw}" must be a positive integer.'})
            continue

        # B. StageNames
        if not stage_name:
            errors.append({'route_code': route_code, 'row': row_num,
                           'message': f'StageNames is empty at stage {stage_order}.'})

        # B. RouteName
        if not route_name:
            errors.append({'route_code': route_code, 'row': row_num, 'message': 'RouteName is empty.'})

        # B. NoOfStages
        no_of_stages_raw = cell(row_values, 'noofstages')
        try:
            no_of_stages = int(float(no_of_stages_raw)) if no_of_stages_raw else 0
            if no_of_stages <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            errors.append({'route_code': route_code, 'row': row_num,
                           'message': f'NoOfStages "{no_of_stages_raw}" must be a positive integer.'})
            no_of_stages = None

        # B. MinFare
        min_fare_raw = cell(row_values, 'minfare')
        min_fare = _to_float(min_fare_raw)
        if min_fare is None:
            errors.append({'route_code': route_code, 'row': row_num,
                           'message': f'MinFare "{min_fare_raw}" must be a non-negative number.'})
        elif min_fare < 0:
            errors.append({'route_code': route_code, 'row': row_num,
                           'message': 'MinFare must be >= 0.'})
            min_fare = None

        # B. FareType
        if fare_type_raw not in ('TABLE', 'GRAPH'):
            errors.append({'route_code': route_code, 'row': row_num,
                           'message': f'FareType "{fare_type_raw}" must be TABLE or GRAPH.'})
            fare_type_int = None
        else:
            fare_type_int = FARE_TYPE_MAP[fare_type_raw]

        # B. Distance
        distance_raw = cell(row_values, 'distance')
        distance = _to_float(distance_raw)
        if distance is None:
            errors.append({'route_code': route_code, 'row': row_num,
                           'message': f'Distance "{distance_raw}" must be a non-negative number.'})
        elif distance < 0:
            errors.append({'route_code': route_code, 'row': row_num, 'message': 'Distance must be >= 0.'})
            distance = None

        # B. Flag columns (0 or 1)
        for flag in FLAG_COLS:
            fv = cell(row_values, flag)
            if fv:
                try:
                    fi = int(float(fv))
                    if fi not in (0, 1):
                        errors.append({'route_code': route_code, 'row': row_num,
                                       'message': f'{flag.title()} must be 0 or 1, got {fi}.'})
                except (ValueError, TypeError):
                    errors.append({'route_code': route_code, 'row': row_num,
                                   'message': f'{flag.title()} must be 0 or 1, got "{fv}".'})

        # B. Fare value
        fare_raw = raw_cell(row_values, fare_cidx)
        fare_val = _to_float(str(fare_raw or ''))
        if fare_val is None:
            errors.append({'route_code': route_code, 'row': row_num,
                           'message': f'Fare value "{fare_raw}" must be a non-negative number.'})
        elif fare_val < 0:
            errors.append({'route_code': route_code, 'row': row_num, 'message': 'Fare value must be >= 0.'})
            fare_val = None

        # GRAPH extra fare columns
        extra_fares = {}  # {origin_stage_k: value}
        if fare_type_raw == 'GRAPH':
            for k, cidx in extra_cols.items():
                v = raw_cell(row_values, cidx)
                if v is not None and str(v).strip():
                    fv = _to_float(str(v))
                    if fv is None or fv < 0:
                        errors.append({'route_code': route_code, 'row': row_num,
                                       'message': f'MatrixCol{k} value "{v}" must be a non-negative number.'})
                    else:
                        extra_fares[k] = fv

        # D. BusType lookup (once per route_code, on first row)
        bus_type_obj = bus_type_map.get(bus_type_raw.lower()) if bus_type_raw else None
        if route_code not in route_groups:
            if not bus_type_obj:
                avail = ', '.join(sorted(bus_type_map.keys())) or 'none'
                errors.append({'route_code': route_code, 'row': row_num,
                               'message': f'BusType "{bus_type_raw}" not found or inactive. Available: {avail}'})

        # Accumulate into route_groups
        if route_code not in route_groups:
            flags = {f: _to_bool(cell(row_values, f)) for f in FLAG_COLS}
            route_groups[route_code] = {
                'route_name':   route_name,
                'bus_type_raw': bus_type_raw,
                'bus_type_obj': bus_type_obj,
                'no_of_stages': no_of_stages,
                'min_fare':     min_fare,
                'fare_type':    fare_type_raw,
                'fare_type_int': fare_type_int,
                'flags':        flags,
                'first_row':    row_num,
                'stages':       [],
            }

        route_groups[route_code]['stages'].append({
            'row':          row_num,
            'stage_order':  stage_order,
            'stage_name':   stage_name,
            'distance':     distance,
            'fare':         fare_val,
            'extra_fares':  extra_fares,
        })

    if not route_groups:
        return None, errors + [{'route_code': None, 'row': None, 'message': 'No valid route data found.'}], warnings, []

    # C. Route-level consistency checks
    for route_code, rg in route_groups.items():
        stages = sorted(rg['stages'], key=lambda s: s['stage_order'])
        n = rg['no_of_stages']

        # C. Row count == NoOfStages
        if n is not None and len(stages) != n:
            errors.append({'route_code': route_code, 'row': rg['first_row'],
                           'message': f'Expected {n} stages but found {len(stages)} data rows.'})

        # C. StageOrder must be exactly 1..N
        orders = [s['stage_order'] for s in stages]
        expected = list(range(1, len(stages) + 1))
        if orders != expected:
            errors.append({'route_code': route_code, 'row': rg['first_row'],
                           'message': f'StageOrder sequence {orders} must be exactly 1..{len(stages)} with no gaps.'})

        # C. Stage 1 fare must be 0
        s1 = next((s for s in stages if s['stage_order'] == 1), None)
        if s1 and s1['fare'] not in (0, 0.0, None):
            errors.append({'route_code': route_code, 'row': s1['row'],
                           'message': f'Stage 1 must have Fare = 0, got {s1["fare"]}.'})

        # C. Distances non-decreasing (warning)
        dists = [s['distance'] for s in stages if s['distance'] is not None]
        for i in range(1, len(dists)):
            if dists[i] < dists[i - 1]:
                warnings.append({'message': f'{route_code}: Distances are not non-decreasing '
                                             f'(stage {i} → {dists[i-1]}, stage {i+1} → {dists[i]}).'})
                break

        # C. GRAPH: stage i should have i-1 total fare values
        if rg['fare_type'] == 'GRAPH':
            for s in stages:
                i = s['stage_order']
                if i == 1:
                    continue
                # Fare col counts as 1, each extra_fares entry adds 1
                actual = 1 + len(s['extra_fares'])  # Fare col + extras
                expected_total = i - 1
                if actual < expected_total:
                    warnings.append({'message': f'{route_code} stage {i}: expected {expected_total} fare '
                                                f'values but found {actual}. Missing fares will default to 0.'})

        # D. Soft-deleted stages
        stage_names = [s['stage_name'] for s in stages if s['stage_name']]
        deleted_stages = set(
            Stage.objects.filter(
                company=company,
                stage_code__in=stage_names,
                is_deleted=True
            ).values_list('stage_code', flat=True)
        )
        for sname in deleted_stages:
            errors.append({'route_code': route_code, 'row': None,
                           'message': f'Stage "{sname}" exists but is soft-deleted. Restore it before importing.'})

    # E. Duplicate check
    all_codes = list(route_groups.keys())
    existing_codes = list(
        Route.objects.filter(
            company=company,
            route_code__in=all_codes,
            is_deleted=False
        ).values_list('route_code', flat=True)
    )

    return route_groups, errors, warnings, existing_codes


def _build_preview(route_groups):
    preview = []
    for rc, rg in route_groups.items():
        flags = rg.get('flags', {})
        concessions = [k for k, v in flags.items() if v]
        preview.append({
            'route_code':   rc,
            'route_name':   rg['route_name'],
            'bus_type':     rg['bus_type_raw'],
            'no_of_stages': rg['no_of_stages'] or len(rg['stages']),
            'fare_type':    rg['fare_type'],
            'concessions':  concessions,
        })
    return preview


def _execute_import(route_groups, company, user, skip_duplicates):
    existing_codes = set(
        Route.objects.filter(
            company=company,
            route_code__in=list(route_groups.keys()),
            is_deleted=False
        ).values_list('route_code', flat=True)
    )

    imported_count = 0
    skipped_count  = 0
    stages_created = 0

    with transaction.atomic():
        for route_code, rg in route_groups.items():

            if route_code in existing_codes:
                if skip_duplicates:
                    skipped_count += 1
                    continue
                else:
                    raise ValueError(f'Route "{route_code}" already exists.')

            bus_type_obj = rg['bus_type_obj']
            if not bus_type_obj:
                raise ValueError(f'BusType for route "{route_code}" not found.')

            flags = rg.get('flags', {})
            route = Route.objects.create(
                route_code=route_code,
                route_name=rg['route_name'],
                min_fare=rg['min_fare'] or 0,
                fare_type=rg['fare_type_int'],
                bus_type=bus_type_obj,
                half=flags.get('half', False),
                luggage=flags.get('luggage', False),
                student=flags.get('student', False),
                adjust=flags.get('adjust', False),
                conc=flags.get('concession', False),
                ph=flags.get('ph', False),
                pass_allow=flags.get('pass', False),
                use_stop=False,
                start_from=0,
                is_deleted=False,
                company=company,
                created_by=user,
                updated_by=user,
            )

            RouteBusType.objects.create(
                route=route, bus_type=bus_type_obj,
                company=company, created_by=user,
            )

            stages_sorted = sorted(rg['stages'], key=lambda s: s['stage_order'])

            for s_data in stages_sorted:
                sname = s_data['stage_name']
                if not sname:
                    continue

                # Resolve or create stage
                existing_stage = Stage.objects.filter(
                    company=company, stage_code=sname, is_deleted=False
                ).first()

                if existing_stage:
                    stage_obj = existing_stage
                else:
                    stage_obj = Stage.objects.create(
                        stage_code=sname,
                        stage_name=sname,
                        company=company,
                        created_by=user,
                    )
                    stages_created += 1

                RouteStage.objects.create(
                    route=route,
                    stage=stage_obj,
                    sequence_no=s_data['stage_order'],
                    distance=s_data['distance'] or 0,
                    company=company,
                    created_by=user,
                )

            # Build fare records
            fares_to_create = []
            if rg['fare_type_int'] == 1:  # TABLE
                # Convention: (row=1, col=stage_i) for each non-zero stage fare
                for s_data in stages_sorted:
                    i = s_data['stage_order']
                    f = s_data['fare']
                    if f is None or f == 0:
                        continue
                    fares_to_create.append(Fare(
                        route=route,
                        row=1, col=i,
                        fare_amount=round(f),
                        route_name=route.route_name,
                        company=company,
                        created_by=user,
                    ))

            else:  # GRAPH
                # Convention: (row=from_stage, col=to_stage-1)
                # Excel: for stage i, Fare col = fare(1→i), MatrixColK = fare(K→i)
                for s_data in stages_sorted:
                    i = s_data['stage_order']
                    if i == 1:
                        continue  # stage 1 row is diagonal only (fare=0)

                    # Fare column → fare from stage 1 to stage i
                    f_main = s_data['fare']
                    if f_main and f_main != 0:
                        fares_to_create.append(Fare(
                            route=route,
                            row=1, col=i - 1,
                            fare_amount=round(f_main),
                            route_name=route.route_name,
                            company=company,
                            created_by=user,
                        ))

                    # MatrixColK → fare from stage K to stage i
                    for k, fv in s_data.get('extra_fares', {}).items():
                        if fv == 0:
                            continue
                        fares_to_create.append(Fare(
                            route=route,
                            row=k, col=i - 1,
                            fare_amount=round(fv),
                            route_name=route.route_name,
                            company=company,
                            created_by=user,
                        ))

            if fares_to_create:
                Fare.objects.bulk_create(fares_to_create, ignore_conflicts=True)

            imported_count += 1

    return imported_count, skipped_count, stages_created


# ── Views ─────────────────────────────────────────────────────────────────────

class RouteImportValidateView(APIView):
    """
    POST /masterdata/routes/import/validate
    Accepts .xlsx file, runs full validation, returns errors/warnings/preview.
    Does NOT write to DB.
    """
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        user, company, err = _auth(request)
        if err:
            return err

        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({'message': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)
        if not excel_file.name.lower().endswith('.xlsx'):
            return Response({'message': 'Only .xlsx files are accepted.'}, status=status.HTTP_400_BAD_REQUEST)

        ws, load_err = _load_worksheet(excel_file.read())
        if load_err:
            return Response({'message': load_err}, status=status.HTTP_400_BAD_REQUEST)

        route_groups, errors, warnings, duplicate_codes = _parse_and_validate(ws, company)
        routes_preview = _build_preview(route_groups) if route_groups else []

        return Response({
            'errors':           errors,
            'warnings':         warnings,
            'duplicate_codes':  duplicate_codes,
            'routes_preview':   routes_preview,
        }, status=status.HTTP_200_OK)


class RouteImportConfirmView(APIView):
    """
    POST /masterdata/routes/import/confirm
    Accepts .xlsx file + skip_duplicates flag, runs import inside transaction.atomic().
    """
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        user, company, err = _auth(request)
        if err:
            return err

        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({'message': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        skip_duplicates_raw = request.data.get('skip_duplicates', 'true')
        skip_duplicates = str(skip_duplicates_raw).lower() in ('true', '1', 'yes')

        ws, load_err = _load_worksheet(excel_file.read())
        if load_err:
            return Response({'message': load_err}, status=status.HTTP_400_BAD_REQUEST)

        route_groups, errors, warnings, duplicate_codes = _parse_and_validate(ws, company)

        # Block hard errors
        if errors:
            return Response({
                'message': 'Import blocked due to validation errors.',
                'errors': errors,
            }, status=status.HTTP_400_BAD_REQUEST)

        if not route_groups:
            return Response({'message': 'No valid routes found in file.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            imported, skipped, stages_created = _execute_import(route_groups, company, user, skip_duplicates)
        except Exception as e:
            return Response({'message': f'Import failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            'message': f'{imported} route(s) imported successfully.',
            'imported_count': imported,
            'skipped_count':  skipped,
            'stages_created': stages_created,
        }, status=status.HTTP_200_OK)


class RouteImportTemplateView(APIView):
    """
    GET /masterdata/routes/import/template/<fare_type>
    Returns a sample .xlsx template for TABLE or GRAPH fare.
    """

    BASE_COLS = [
        'RouteName', 'RouteCode', 'BusType', 'NoOfStages', 'MinFare',
        'FareType', 'Distance', 'Half', 'Student', 'Adjust', 'Luggage',
        'PH', 'Concession', 'Pass', 'Ladies', 'Senior', 'StageOrder', 'StageNames', 'Fare',
    ]

    # Placeholder rows: show expected format without real data
    # Flags: 0 or 1 | FareType: TABLE | BusType: must match active bustype_code
    TABLE_PLACEHOLDERS = [
        ['<RouteName>', '<CODE>', '<BusTypeCode>', '<N>', '<MinFare>', 'TABLE', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, '<StageName1>', 0],
        ['<RouteName>', '<CODE>', '<BusTypeCode>', '<N>', '<MinFare>', 'TABLE', '<km>', 0, 0, 0, 0, 0, 0, 0, 0, 0, 2, '<StageName2>', '<Fare>'],
        ['<RouteName>', '<CODE>', '<BusTypeCode>', '<N>', '<MinFare>', 'TABLE', '<km>', 0, 0, 0, 0, 0, 0, 0, 0, 0, 3, '<StageName3>', '<Fare>'],
    ]

    GRAPH_COLS = BASE_COLS + ['MatrixCol2', 'MatrixCol3']

    GRAPH_PLACEHOLDERS = [
        ['<RouteName>', '<CODE>', '<BusTypeCode>', '<N>', '<MinFare>', 'GRAPH', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, '<StageName1>', 0, '', ''],
        ['<RouteName>', '<CODE>', '<BusTypeCode>', '<N>', '<MinFare>', 'GRAPH', '<km>', 0, 0, 0, 0, 0, 0, 0, 0, 0, 2, '<StageName2>', '<Fare1to2>', '', ''],
        ['<RouteName>', '<CODE>', '<BusTypeCode>', '<N>', '<MinFare>', 'GRAPH', '<km>', 0, 0, 0, 0, 0, 0, 0, 0, 0, 3, '<StageName3>', '<Fare1to3>', '<Fare2to3>', ''],
    ]

    def get(self, request, fare_type):
        fare_type = fare_type.upper()
        if fare_type not in ('TABLE', 'GRAPH'):
            return Response({'message': 'fare_type must be TABLE or GRAPH.'}, status=status.HTTP_400_BAD_REQUEST)

        wb = Workbook()
        ws = wb.create_sheet('SRE_Import')
        del wb[wb.sheetnames[0]]  # remove default Sheet

        if fare_type == 'TABLE':
            headers = self.BASE_COLS
            sample_rows = self.TABLE_PLACEHOLDERS
        else:
            headers = self.GRAPH_COLS
            sample_rows = self.GRAPH_PLACEHOLDERS

        ws.append(headers)
        for row in sample_rows:
            ws.append(row)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'SRE_Import_{fare_type}.xlsx'
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
